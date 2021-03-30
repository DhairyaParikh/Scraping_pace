import os
from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.storage import default_storage, FileSystemStorage
import _io
import io
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup as bs
import requests
import matplotlib.pyplot as plt
import openpyxl as pxl
from selenium.webdriver.common.keys import Keys
import time
import sys
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from datetime import date
from selenium import webdriver
import pandas as pd
sys.path.insert(0,'/usr/lib/chromium-browser/chromedriver')
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
header = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.75 Safari/537.36",  "X-Requested-With": "XMLHttpRequest"}

#start  call all the program
def start(request):
    fs = FileSystemStorage()
    # df = pd.read_excel (r'media/PaceKeyworks.xlsx', sheet_name='Cybersecurity_Skill')
    # df = pd.read_excel(fs.open("Scraping_pace/media/PaceKeywords.xlsx", "r"))
    # xls = xlrd.open_workbook(r'Scraping_pace/media/PaceKeywords.xlsx', on_demand=True)
    wb = load_workbook(filename='PaceKeywords.xlsx', read_only=True)

    return

def loadSheet(sheet_name):
    print(sheet_name)
    search_terms = pd.read_excel('PaceKeywords.xlsx',sheet_name=sheet_name)
    return search_terms


#calling all the scraping prigram
def scraping(request):
    # wb = load_workbook(filename='PaceKeywords.xlsx', read_only=True)

    search_terms_skills = loadSheet("Software_Skill")
    search_terms_positions = loadSheet("Software_Positions")
    search_terms_BC = loadSheet("Software_BadgeAndCert")

    indeed_Skills_listOflist = indeed_analysis("", search_terms_skills)
    indeed_Skills_list = indeed_Skills_listOflist[0]
    indeed_Skills_list_url = indeed_Skills_listOflist[1]

    monster_skills_listOflist = monster_analysis("", search_terms_skills)
    monster_skills_list = monster_skills_listOflist[0]
    monster_skills_list_url = monster_skills_listOflist[1]

    careerbuilderjobs_skills_listOflist = careerbuilderjobs_analysis("", search_terms_skills)
    careerbuilderjobs_skills_list = careerbuilderjobs_skills_listOflist[0]
    careerbuilderjobs_skills_list_url = careerbuilderjobs_skills_listOflist[1]


    flexjobs_Skill_listOflist = flexjobs_analysis("", search_terms_skills)
    flexjobs_Skill_list = flexjobs_Skill_listOflist[0]
    flexjobs_Skill_list_url = flexjobs_Skill_listOflist[1]

    simply_skills_listOflist = simplyhired_analysis("", search_terms_skills)
    simply_skills_list = simply_skills_listOflist[0]
    simply_skills_list_url = simply_skills_listOflist[1]

    gartner_skills_listOflist = gartner_analysis("", search_terms_skills)
    gartner_skills_list = gartner_skills_listOflist[0]
    gartner_skills_list_url = gartner_skills_listOflist[1]

    dice_skills_listOflist = dice_analysis("", search_terms_skills)
    dice_skills_list = dice_skills_listOflist[0]
    dice_skills_list_url = dice_skills_listOflist[1]

    findjobs_skills_listOflist = findjobs_analysis("", search_terms_skills)
    findjobs_skills_list = findjobs_skills_listOflist[0]
    findjobs_skills_list_url = findjobs_skills_listOflist[1]

    df_skills = pd.DataFrame(indeed_Skills_list, index=search_terms_skills.iloc[:, 0], columns=['Indeed'])
    df_skills_url = pd.DataFrame(indeed_Skills_list_url, index=search_terms_skills.iloc[:, 0], columns=['Indeed'])

    df_skills.rename(index={'': 'CS_Skill'}, inplace=True)
    df_skills_url.rename(index={'': 'CS_Skill'}, inplace=True)

    df_skills['Gartner'] = gartner_skills_list
    df_skills['Dice'] = dice_skills_list
    df_skills['Monster'] = monster_skills_list
    df_skills['SimplyHired'] = simply_skills_list
    # df_skills['Glassdoor'] = glassdoor_skills_list
    df_skills['flexjobs'] = flexjobs_Skill_list
    df_skills['CareerBuilder'] = careerbuilderjobs_skills_list
    df_skills['findJobs'] = findjobs_skills_list
    df_skills['Date'] = date.today()
    new_order = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8]
    df_skills = df_skills[df_skills.columns[new_order]]
    with pd.ExcelWriter('PortalsCountsForCyberSec.xlsx') as writer:
        df_skills.to_excel(writer, sheet_name='SkillsAnalysis')

    df_skills_url['Gartner'] = gartner_skills_list_url
    df_skills_url['Dice'] = dice_skills_list_url
    df_skills_url['Monster'] = monster_skills_list_url
    df_skills_url['SimplyHired'] = simply_skills_list_url
    # df_skills_url['Glassdoor'] = glassdoor_skills_list
    df_skills_url['flexjobs'] = flexjobs_Skill_list_url
    df_skills_url['CareerBuilder'] = careerbuilderjobs_skills_list_url
    df_skills_url['findJobs'] = findjobs_skills_list_url
    df_skills_url['Date'] = date.today()
    df_skills_url = df_skills_url[df_skills_url.columns[new_order]]

    with pd.ExcelWriter('PortalsCountsForCyberSecURL.xlsx') as writer:
        df_skills_url.to_excel(writer, sheet_name='SkillsAnalysisURL')
    df_skills_r = None
    Ct = 1

    for role in search_terms_positions.iloc[0:3, 0]:
        roleFinal = role
        print(roleFinal)
        indeed_Skills_listOflist = indeed_analysis(roleFinal, search_terms_skills)
        indeed_Skills_list1 = indeed_Skills_listOflist[0]
        indeed_Skills_list_url1 = indeed_Skills_listOflist[1]

        monster_skills_listOflist = monster_analysis(roleFinal, search_terms_skills)
        monster_skills_list1 = monster_skills_listOflist[0]
        monster_skills_list_url1 = monster_skills_listOflist[1]

        careerbuilderjobs_skills_listOflist = careerbuilderjobs_analysis(roleFinal, search_terms_skills)
        careerbuilderjobs_skills_list1 = careerbuilderjobs_skills_listOflist[0]
        careerbuilderjobs_skills_list_url1 = careerbuilderjobs_skills_listOflist[1]

        flexjobs_Skill_listOflist = flexjobs_analysis(roleFinal, search_terms_skills)
        flexjobs_Skill_list1 = flexjobs_Skill_listOflist[0]
        flexjobs_Skill_list_url1 = flexjobs_Skill_listOflist[1]

        simply_skills_listOflist = simplyhired_analysis(roleFinal, search_terms_skills)
        simply_skills_list1 = simply_skills_listOflist[0]
        simply_skills_list_url1 = simply_skills_listOflist[1]

        dice_skills_listOflist = dice_analysis(roleFinal, search_terms_skills)
        dice_skills_list1 = dice_skills_listOflist[0]
        dice_skills_list_url1 = dice_skills_listOflist[1]

        findjobs_skills_listOflist = findjobs_analysis(roleFinal, search_terms_skills)
        findjobs_skills_list1 = findjobs_skills_listOflist[0]
        findjobs_skills_list_url1 = findjobs_skills_listOflist[1]

        # if df_skills_r is None:
        df_skills_r = pd.DataFrame(indeed_Skills_list1, index=search_terms_skills.iloc[:, 0], columns=['Indeed'])
        df_skills_r.rename(index={'': 'Data_Skill_r'}, inplace=True)
        # GARTNER Double Qutoes with postion and skill does not work . df_skills_r['Gartner'] = gartner_skills_list1
        df_skills_r['Dice'] = dice_skills_list1
        df_skills_r['Monster'] = monster_skills_list1
        df_skills_r['SimplyHired'] = simply_skills_list1
        # df_skills_r['Glassdoor'] = glassdoor_skills_list1
        df_skills_r['flexjobs'] = flexjobs_Skill_list1
        df_skills_r['CareerBuilder'] = careerbuilderjobs_skills_list1
        df_skills_r['findJobs'] = findjobs_skills_list1
        df_skills_r['Date'] = date.today()
        new_order = [-1, 0, 1, 2, 3, 4, 5, 6, 7]
        df_skills_r = df_skills_r[df_skills_r.columns[new_order]]

        df_skills_r_url = pd.DataFrame(indeed_Skills_list_url1, index=search_terms_skills.iloc[:, 0],
                                       columns=['Indeed'])
        df_skills_r_url.rename(index={'': 'Skill_r'}, inplace=True)
        # GARTNER Double Qutoes with postion and skill does not work .df_skills_url['Gartner'] = gartner_skills_list_url1
        df_skills_r_url['Dice'] = dice_skills_list_url1
        df_skills_r_url['Monster'] = monster_skills_list_url1
        df_skills_r_url['SimplyHired'] = simply_skills_list_url1
        # df_skills_url['Glassdoor'] = glassdoor_skills_list
        df_skills_r_url['flexjobs'] = flexjobs_Skill_list_url1
        df_skills_r_url['CareerBuilder'] = careerbuilderjobs_skills_list_url1
        df_skills_r_url['findJobs'] = findjobs_skills_list_url1
        df_skills_r_url['Date'] = date.today()
        df_skills_r_url = df_skills_r_url[df_skills_r_url.columns[new_order]]
        with pd.ExcelWriter('PortalsCountsForCyberSec.xlsx', mode='a') as writer:
            df_skills_r.to_excel(writer, sheet_name=roleFinal + '_S')
        with pd.ExcelWriter('PortalsCountsForCyberSecURL.xlsx', mode='a') as writer:
            df_skills_r_url.to_excel(writer, sheet_name=roleFinal + '_URL')
        Ct = Ct + 1

    careerbuilderjobs_positions_listOflist = careerbuilderjobs_analysis("", search_terms_positions)
    careerbuilderjobs_positions_list = careerbuilderjobs_positions_listOflist[0]
    careerbuilderjobs_positions_list_url = careerbuilderjobs_positions_listOflist[1]

    flexjobs_positions_listOflist = flexjobs_analysis("", search_terms_positions)
    flexjobs_positions_list = flexjobs_positions_listOflist[0]
    flexjobs_positions_list_url = flexjobs_positions_listOflist[1]

    # glassdoor_positions_list = glassdoor_analysis("Data",search_terms_positions)
    findJobs_positions_listOflist = findjobs_analysis("", search_terms_positions)
    findjobs_positions_list = findJobs_positions_listOflist[0]
    findjobs_positions_list_url = findJobs_positions_listOflist[1]

    gartner_positions_listOflist = gartner_analysis("", search_terms_positions)
    gartner_positions_list = gartner_positions_listOflist[0]
    gartner_positions_list_url = gartner_positions_listOflist[1]

    indeed_positions_listOflist = indeed_analysis("", search_terms_positions)
    indeed_positions_list = indeed_positions_listOflist[0]
    indeed_positions_list_url = indeed_positions_listOflist[1]

    monster_positions_listOflist = monster_analysis("", search_terms_positions)
    monster_positions_list = monster_positions_listOflist[0]
    monster_positions_list_url = monster_positions_listOflist[1]

    careerbuilderjobs_positions_listOflist = careerbuilderjobs_analysis("", search_terms_positions)
    careerbuilderjobs_positions_list = careerbuilderjobs_positions_listOflist[0]
    careerbuilderjobs_positions_list_url = careerbuilderjobs_positions_listOflist[1]

    simply_positions_listOflist = simplyhired_analysis("", search_terms_positions)
    simply_positions_list = simply_positions_listOflist[0]
    simply_positions_list_url = simply_positions_listOflist[1]

    dice_positions_listOflist = dice_analysis("", search_terms_positions)
    dice_positions_list = dice_positions_listOflist[0]
    dice_positions_list_url = dice_positions_listOflist[1]

    findjobs_positions_listOflist = findjobs_analysis("", search_terms_positions)
    findjobs_positions_list = findjobs_positions_listOflist[0]
    findjobs_positions_list_url = findjobs_positions_listOflist[1]

    df_positions = pd.DataFrame(indeed_positions_list, index=search_terms_positions.iloc[:, 0], columns=['Indeed'])
    df_positions.rename(index={'': 'Data_positions'}, inplace=True)

    df_positions_url = pd.DataFrame(indeed_positions_list_url, index=search_terms_positions.iloc[:, 0],
                                    columns=['Indeed'])
    df_positions_url.rename(index={'': 'Data_positions url'}, inplace=True)

    df_positions['Gartner'] = gartner_positions_list
    df_positions['Dice'] = dice_positions_list
    df_positions['Monster'] = monster_positions_list
    df_positions['SimplyHired'] = simply_positions_list
    # df_positions['Glassdoor'] = glassdoor_positions_list
    df_positions['flexjobs'] = flexjobs_positions_list
    df_positions['CareerBuilder'] = careerbuilderjobs_positions_list
    df_positions['findJobs'] = findjobs_positions_list
    df_positions['Date'] = date.today()
    new_order = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8]
    df_positions = df_positions[df_positions.columns[new_order]]

    df_positions_url['Gartner'] = gartner_positions_list_url
    df_positions_url['Dice'] = dice_positions_list_url
    df_positions_url['Monster'] = monster_positions_list_url
    df_positions_url['SimplyHired'] = simply_positions_list_url
    # df_positions_url['Glassdoor'] = glassdoor_positions_list
    df_positions_url['flexjobs'] = flexjobs_positions_list_url
    df_positions_url['CareerBuilder'] = careerbuilderjobs_positions_list_url
    df_positions_url['Usjobs'] = findjobs_positions_list_url
    df_positions_url['Date'] = date.today()
    df_positions_url = df_positions_url[df_positions_url.columns[new_order]]

    with pd.ExcelWriter('PortalsCountsForCyberSec.xlsx', mode='a') as writer:
        df_positions.to_excel(writer, sheet_name='CyberSecurityPositionsAnalysis')
    with pd.ExcelWriter('PortalsCountsForCyberSecURL.xlsx', mode='a') as writer:
        df_positions_url.to_excel(writer, sheet_name='CyberSecurityPositionsAnalysisUrl')

    careerbuilder_BC_listOflist = careerbuilderjobs_analysis("", search_terms_BC)
    careerbuilder_BC_list = careerbuilder_BC_listOflist[0]
    careerbuilder_BC_list_url = careerbuilder_BC_listOflist[1]

    flexjobs_BC_listOflist = flexjobs_analysis("", search_terms_BC)
    flexjobs_BC_list = flexjobs_BC_listOflist[0]
    flexjobs_BC_list_url = flexjobs_BC_listOflist[1]

    indeed_BC_listOflist = indeed_analysis("", search_terms_BC)
    indeed_BC_list = indeed_BC_listOflist[0]
    indeed_BC_list_url = indeed_BC_listOflist[1]

    indeed_BC_listOflist = indeed_analysis("", search_terms_BC)
    indeed_BC_list = indeed_BC_listOflist[0]
    indeed_BC_list_url = indeed_BC_listOflist[1]

    monster_BC_listOflist = monster_analysis("", search_terms_BC)
    monster_BC_list = monster_BC_listOflist[0]
    monster_BC_list_url = monster_BC_listOflist[1]

    gartner_BC_listOflist = gartner_analysis("", search_terms_BC)
    gartner_BC_list = gartner_BC_listOflist[0]
    gartner_BC_list_url = gartner_BC_listOflist[1]

    simply_BC_listOflist = simplyhired_analysis("", search_terms_BC)
    simply_BC_list = simply_BC_listOflist[0]
    simply_BC_list_url = simply_BC_listOflist[1]

    dice_BC_listOflist = dice_analysis("", search_terms_BC)
    dice_BC_list = dice_BC_listOflist[0]
    dice_BC_list_url = dice_BC_listOflist[1]

    findjobs_BC_listOflist = findjobs_analysis("", search_terms_BC)
    findjobs_BC_list = findjobs_BC_listOflist[0]
    findjobs_BC_list_url = findjobs_BC_listOflist[1]

    df_BC = pd.DataFrame(indeed_BC_list, index=search_terms_BC.iloc[:, 0], columns=['Indeed '])

    df_BC.rename(index={'': 'Data_BC'}, inplace=True)

    df_BC_url = pd.DataFrame(indeed_BC_list_url, index=search_terms_BC.iloc[:, 0], columns=['Indeed '])

    df_BC_url.rename(index={'': 'Data_BC_url'}, inplace=True)

    df_BC['Dice'] = dice_BC_list
    df_BC['Gartner'] = gartner_BC_list
    df_BC['Monster'] = monster_BC_list
    df_BC['SimplyHired'] = simply_BC_list
    # df_BC['Glassdoor'] = glassdoor_BC_list
    df_BC['flexjobs'] = flexjobs_BC_list
    df_BC['CareerBuilder'] = careerbuilder_BC_list
    df_BC['findJobs'] = findjobs_BC_list
    df_BC['Date'] = date.today()
    new_order = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8]
    df_BC = df_BC[df_BC.columns[new_order]]

    df_BC_url['Gartner'] = gartner_BC_list_url
    df_BC_url['Dice'] = dice_BC_list_url
    df_BC_url['Monster'] = monster_BC_list_url
    df_BC_url['SimplyHired'] = simply_BC_list_url
    # df_BC_url['Glassdoor'] = glassdoor_BC_list
    df_BC_url['flexjobs'] = flexjobs_BC_list_url
    df_BC_url['CareerBuilder'] = careerbuilder_BC_list_url
    df_BC_url['findJobs'] = findjobs_BC_list_url
    df_BC_url['Date'] = date.today()
    df_BC_url = df_BC_url[df_BC_url.columns[new_order]]

    with pd.ExcelWriter('PortalsCountsForCyberSec.xlsx', mode='a') as writer:
        df_BC.to_excel(writer, sheet_name='CyberSecurityBCAnalysis')

    with pd.ExcelWriter('PortalsCountsForCyberSecURL.xlsx', mode='a') as writer:
        df_BC_url.to_excel(writer, sheet_name='CyberSecurityBCAnalysisUrl')
    return HttpResponse('200')



def indeed_analysis(sheet_name, search_terms_p):
    search_terms = search_terms_p
    indeed_list = []
    indeed_list_urls = []

    search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')

    for term in search_terms.iloc[:, 0]:
        term = term.replace(" ", "%20")
        term = "\"" + term + "\""
        if sheet_name == "":
            url = f'https://www.indeed.com/jobs?q={term}&l=United%20States&vjk=ea8b83a76166217a'
        else:
            sheet_namech = sheet_name.replace(" ", "%20")
            sheet_namech = "\"" + sheet_namech + "\""
            url = f'https://www.indeed.com/jobs?q={sheet_namech}%20{term}&l=United%20States&vjk=ea8b83a76166217a'
        indeed_list_urls.append(url)
        try:
            r = requests.get(url, headers=header, timeout=10)
            soup = bs(r.text, 'html.parser')
            count_str = soup.find('div', id="searchCountPages").get_text()
            numb = count_str.split()
            indeed_list.append(numb[-2])
        except Exception as e:
            # print(f'error: {e}')
            indeed_list.append(0)
    return [indeed_list, indeed_list_urls]

def monster_analysis(sheet_name,search_terms_p):
  monster_list = []
  monster_list_urls = []
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  search_terms.iloc[:, 0] = search_terms.replace(' ', '-', regex=True)

  for term in search_terms.iloc[:, 0]:
      #print(term)
      term = term.replace(" ","%20")
      #term = "\""+term+"\""
      if sheet_name == "":
        url1 = f'https://www.monster.com/jobs/search/?q=__22{term}__22&where=USA'
        url2 = f'https://www.monster.com/jobs/search/?q=__27{term}__27&where=USA'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        url1 = f'https://www.monster.com/jobs/search/?q=__22{sheet_namech}__22-__22{term}__22&where=USA'
        url2 = f'https://www.monster.com/jobs/search/?q=__27{sheet_namech}__27-__27{term}__27&where=USA'
      monster_list_urls.append(url1)
      try:
          r = requests.get(url1, headers=header, timeout=5)
          soup = bs(r.text, 'html.parser')
          count_str = soup.find('h2', class_="figure").get_text()
        # print(count_str)
          numb = count_str.split()
          monster_count = numb[0].replace("(", "")
          monster_list.append(monster_count)
      except Exception as e:
      #   print(f'error: {e}')
          monster_list.append(0)
  return [monster_list,monster_list_urls]

def simplyhired_analysis(sheet_name,search_terms_p):
  simply_list = []
  simply_list_url = []
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')

  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      term = "\""+term+"\""
      if sheet_name == "":
        url = f'https://www.simplyhired.com/search?q={term}&l=United+States'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        sheet_namech = "\""+sheet_namech+"\""
        url = f'https://www.simplyhired.com/search?q={sheet_namech}+{term}&l=United+States'
      simply_list_url.append(url)
      try:
          r = requests.get(url, headers=header, timeout=5)
          soup = bs(r.text, 'html.parser')
          count_str = soup.find('span', class_="CategoryPath-total").get_text()
          #print(count_str)
          simply_list.append(count_str.replace(',',''))
      except Exception as e:
          #print(f'error: {e}')
          simply_list.append(0)
  return [simply_list,simply_list_url]

def gartner_analysis(sheet_name,search_terms_p):
  gartner_list = []
  gartner_list_url = []
  search_terms=search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')

  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      if sheet_name == "":
        url = f'https://jobs.gartner.com/search-jobs/{term}/United%20States/494/1/2/6252001/39x7599983215332/-98x5/50/2'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        url = f'https://jobs.gartner.com/search-jobs/{sheet_namech}%20{term}/United%20States/494/1/2/6252001/39x7599983215332/-98x5/50/2'
      gartner_list_url.append(url)
      try:
          wd = webdriver.Chrome('chromedriver',options=chrome_options)
          r = wd.get(url)
          #print(r)
          count_str = (wd.find_elements_by_class_name("results-count-heading"))[0].text
          #print(count_str)
          numb = count_str.split()
          gartner_list.append(numb[-2])
      except Exception as e:
          #print(f'error: {e}')
          gartner_list.append(0)
  return [gartner_list,gartner_list_url]

def dice_analysis(sheet_name,search_terms_p):
  dice_list = []
  dice_list_url = []
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  #print("test")
  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      term = "\""+term+"\""
      if sheet_name == "":
          url = f'https://www.dice.com/jobs?q={term}&location=United%20States'
      else:
         sheet_namech = sheet_name.replace(" ","%20")
         sheet_namech = "\""+sheet_namech+"\""
         url = f'https://www.dice.com/jobs?q={sheet_namech}%20{term}&location=United%20States'
      dice_list_url.append(url)
      timeout = 5
      try:
       wd = webdriver.Chrome('chromedriver',options=chrome_options)
       wd.get(url)
       element_present = EC.presence_of_element_located((By.ID, 'totalJobCount'))
       WebDriverWait(wd, timeout).until(element_present)
       count_str = (wd.find_element_by_id("totalJobCount")).get_attribute("innerText")
       numb = count_str.replace(",", "");
       dice_list.append(numb)
      except Exception as e:
        #print(f'error: {e}')
        dice_list.append(0)
  return [dice_list,dice_list_url]

def glassdoor_analysis(sheet_name,search_terms_p):
  glassdoor_list = []
  glassdoor_list_url = []
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  #print("test")
  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","+")
      term = "\""+term+"\""
      if sheet_name == "":
          #url = f'https://www.glassdoor.com/Job/jobs.htm?suggestCount=0&suggestChosen=false&clickSource=searchBtn&typedKeyword=%22{term}%22&locT=N&locId=1&jobType=&context=Jobs&sc.keyword=%22{term}%22&dropdown=0
          url = f'https://www.glassdoor.com/Job/jobs.htm?suggestCount=0&suggestChosen=false&clickSource=searchBtn&typedKeyword={term}&sc.keyword={term}&locT=N&locId=1'
      else:
         sheet_namech = sheet_name.replace(" ","+")
         sheet_namech = "\""+sheet_namech+"\""
         url = f'https://www.glassdoor.com/Job/jobs.htm?suggestCount=0&suggestChosen=false&clickSource=searchBtn&typedKeyword={sheet_namech}+{term}&sc.keyword={sheet_namech}+{term}&locT=N&locId=1'
      glassdoor_list_url.append(url)
      timeout = 5
      try:
        wd = webdriver.Chrome('chromedriver',options=chrome_options)
        wd.get(url)
        count_str = (wd.find_element_by_xpath("//div[@class='hideHH css-19rczgc e15r6eig0']")).get_attribute("innerText")
        numb_split = count_str.split()
        numb = numb_split[0].replace(",", "")
        glassdoor_list.append(numb)
      except Exception as e:
        #print(f'error: {e}')
        glassdoor_list.append(0)
  return [glassdoor_list,glassdoor_list_url]

def careerbuilderjobs_analysis(sheet_name,search_terms_p):
  careerbuilderjobs_list = []
  careerbuilderjobs_list_url = []
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  #print("test")
  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      term = "\""+term+"\""
      if sheet_name == "":
        url = f'https://www.careerbuilder.com/jobs?utf8=%E2%9C%93&keywords={term}&location=United+States'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        sheet_namech = "\""+sheet_namech+"\""
        url = f'https://www.careerbuilder.com/jobs?utf8=%E2%9C%93&keywords={sheet_namech}+{term}&location=United+States'
      careerbuilderjobs_list_url.append(url)
      timeout = 5
      try:
        wd = webdriver.Chrome('chromedriver',options=chrome_options)
        wd.get(url)
        count_str = (wd.find_element_by_xpath("//h1[@class='fz1rem']")).get_attribute("innerText")
        numb_split = count_str.split()
        for i in numb_split:
          numb = i.replace(",", "")
          if numb.isdigit():
            careerbuilderjobs_list.append(numb)
            break
      except Exception as e:
        #print(f'error: {e}')
        careerbuilderjobs_list.append(0)
  return [careerbuilderjobs_list,careerbuilderjobs_list_url]

def flexjobs_analysis(sheet_name,search_terms_p):
  flexjobs_list = []
  flexjobs_list_url=[]
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  #print("test")
  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      term = "\""+term+"\""
      if sheet_name == "":
        url = f'https://www.flexjobs.com/search?search=&search={term}&location=United+States'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        sheet_namech = "\""+sheet_namech+"\""
        url = f'https://www.flexjobs.com/search?search=&search={sheet_namech}+{term}&location=United+States'
      flexjobs_list_url.append(url)
      timeout = 5
      try:
        wd = webdriver.Chrome('chromedriver',options=chrome_options)
        wd.get(url)
        count_str = (wd.find_element_by_xpath("//h4[@style='margin:0;font-size:14px;']")).get_attribute("innerText")
        numb_split = count_str.split("of")[1].split("for")
        numb = numb_split[0].replace(",", "")
        flexjobs_list.append(numb)
      except Exception as e:
        #print(f'error: {e}')
        flexjobs_list.append(0)
  return [flexjobs_list,flexjobs_list_url]

def findjobs_analysis(sheet_name,search_terms_p):
  usjobs_list = []
  usjobs_list_url =[]
  search_terms = search_terms_p
  search_terms.iloc[:, 0] = search_terms.iloc[:, 0].str.strip('[]')
  #print("test")
  for term in search_terms.iloc[:, 0]:
      term = term.replace(" ","%20")
      term = "\""+term+"\""
      if sheet_name == "":
        url = f'https://find.jobs/jobs-near-me/job-search-results/kw-{term}/co-226/'
      else:
        sheet_namech = sheet_name.replace(" ","%20")
        sheet_namech = "\""+sheet_namech+"\""
        url = f'https://find.jobs/jobs-near-me/job-search-results/kw-{sheet_namech}-{term}/co-226/'
      usjobs_list_url.append(url)
      timeout = 5
      try:
        wd = webdriver.Chrome('chromedriver',options=chrome_options)
        wd.get(url)
        count_str = (wd.find_element_by_xpath("//p[@class='results-count']")).get_attribute("innerText")
        numb_split = count_str.split(" ")[3]
        numb = numb_split.replace(",", "")
        usjobs_list.append(numb)
      except Exception as e:
        #print(f'error: {e}')
        usjobs_list.append(0)
  return [usjobs_list,usjobs_list_url]